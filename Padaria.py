from tkinter import *
from tkinter import ttk
import pandas as pd
import sqlite3
import os
import openpyxl
from datetime import datetime
import random
##Criação do banco de dados
conn = sqlite3.connect("Padaria.DB")
conn.execute(''' CREATE TABLE IF NOT EXISTS produtos (
               Nome text,
               Valor integer,
               Quantidade integer,
               Código INTEGER PRIMARY KEY AUTOINCREMENT
               )''')

conn.execute(''' CREATE TABLE IF NOT EXISTS vendas (
               Produtos text,
               Valor integer,
               Quantidade integer,
               Código text,
               Data DATETIME,
               Hora DATETIME)''')
conn.commit()
#conn.close()

# Listas que serão utilizadas
codigo_venda = None
lista_venda = []
lista = list(range(1, 16))

# Criação da janela inicial
janela_inicial = Tk()
background_image = PhotoImage(file="pao.png")
background_image2 = PhotoImage(file="pao.png")
background_image3 = PhotoImage(file="Data-science.png")

#Função para seleção de relatório mensal ou diário
def janela_excel():
  janela_excel = Toplevel(janela_inicial)
  janela_excel.title("Relatórios")
  janela_excel_button1 = Button(janela_excel, text="registro diário", command=lambda:gerar_excel_dias())
  janela_excel_button1.pack(pady=5)
  janela_excel_button2 = Button(janela_excel, text="registro mensal",command=lambda:gerar_excel_mes())
  janela_excel_button2.pack(pady=5)

#Relatório mensal
def gerar_excel_mes():
   
    # Consulta SQL para obter os dados por mês, incluindo a totalização do mês anterior
    dados_por_mes = conn.execute(
        """
        SELECT strftime('%m/%Y', Data) AS Mes, Produtos, 
               SUM(Quantidade) AS Quantidade, 
               SUM(Valor) AS Valor
        FROM vendas 
        GROUP BY Mes, Produtos
        """
    ).fetchall()

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    worksheet['A1'] = 'Mês'
    worksheet['B1'] = 'Produto'
    worksheet['C1'] = 'Quantidade'
    worksheet['D1'] = 'Valor'

    linha_atual = 2
    total_mes_anterior = 0
    for mes, produto, quantidade, valor in dados_por_mes:
        if linha_atual > 2 and mes != mes_anterior:
            linha_atual += 1
            worksheet.cell(row=linha_atual, column=1, value="Total Mês Anterior")
            worksheet.cell(row=linha_atual, column=4, value=total_mes_anterior)
            total_mes_anterior = 0
        
        # Adiciona os dados do produto atual
        worksheet.cell(row=linha_atual, column=1, value=mes)
        worksheet.cell(row=linha_atual, column=2, value=produto)
        worksheet.cell(row=linha_atual, column=3, value=quantidade)
        worksheet.cell(row=linha_atual, column=4, value=valor)
        linha_atual += 1
        
        # Adiciona o valor atual à totalização do mês anterior
        total_mes_anterior += valor
        mes_anterior = mes

    # Adiciona a última linha com a totalização do mês anterior
    linha_atual += 1
    worksheet.cell(row=linha_atual, column=1, value="Total Mês Anterior")
    worksheet.cell(row=linha_atual, column=4, value=total_mes_anterior)

    # Salva o arquivo do Excel
    now = datetime.now()
    nome_arquivo = f"Relatorio_de_Vendas_{now.strftime('%m-%Y')}.xlsx"
    workbook.save(nome_arquivo)
    os.startfile(nome_arquivo)

#Relatório diário
def gerar_excel_dias():
    # Consulta SQL para obter dados por dia, incluindo uma linha extra para somar os valores do dia
    dados_por_dia = conn.execute("""
        SELECT 
            date(Data), 
            Produtos, 
            SUM(Quantidade), 
            SUM(Valor) 
        FROM vendas 
        GROUP BY date(Data), Produtos
        UNION ALL
        SELECT 
            date(Data),
            'TOTAL DO DIA',
            NULL,
            SUM(Valor)
        FROM vendas
        GROUP BY date(Data)
    """).fetchall()
    
    # Cria uma nova planilha do Excel e seleciona a planilha
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Define os títulos das colunas na primeira linha da planilha
    worksheet['A1'] = 'Dia'
    worksheet['B1'] = 'Produto'
    worksheet['C1'] = 'Quantidade'
    worksheet['D1'] = 'Valor'

    # Itera pelos resultados da consulta SQL e adiciona os valores às células correspondentes na planilha
    linha_atual = 2
    for dia, produto, quantidade, valor in dados_por_dia:
        worksheet.cell(row=linha_atual, column=1, value=dia)
        worksheet.cell(row=linha_atual, column=2, value=produto)
        worksheet.cell(row=linha_atual, column=3, value=quantidade)
        worksheet.cell(row=linha_atual, column=4, value=valor)
        linha_atual += 1

    # Salva a planilha em um arquivo Excel e o executa
    workbook.save("Relatorio_de_Vendas.xlsx")
    os.startfile('Relatorio_de_Vendas.xlsx')


#Gera planilha excel a partir dos dados do banco de dados
def consulta_estoque():
   conn = sqlite3.connect('Padaria.DB')
   
   df = pd.read_sql_query("SELECT * FROM produtos", conn)
   df.to_excel('Estoque.xlsx', index=False)
   
   os.startfile('Estoque.xlsx')
   conn.close()

def abrir_janela_venda(janela_inicial):
    # Cria uma nova janela de vendas
    janela_venda = Toplevel(janela_inicial)
    janela_venda.geometry("600x450")
    janela_venda.resizable(False, False)
    janela_venda.title("Tela de venda")

    # Define a imagem de fundo da janela
    background_label = Label(janela_venda, image=background_image)
    background_label.place(x=0, y=0, relwidth=1, relheight=1)

    # Limpa a lista de vendas
    def limpar_lista():
        lista_venda.clear()

    # Salva o produto na lista de vendas ou adiciona manualmente um valor de pesagem
    def guardar_produto():
        global codigo_venda
        codigo = janela_venda_entry1.get()
        if codigo == "0":
            # Cria uma nova janela para inserir o valor manualmente
            manual = Toplevel(janela_venda)
            manual.title("Valor manual")
            manual.geometry("200x100")

            manual_label = Label(manual, text="Insira o valor:")
            manual_label.pack(pady=5)

            manual_entry = Entry(manual)
            manual_entry.pack(pady=5)

            # Adiciona o valor manual à lista de vendas
            manual_button = Button(manual, text="Adicionar", command=lambda: adicionar_pesagem(manual_entry.get()))
            manual_button.pack(pady=5)

            return
        #Pega os dados no banco de dados, verifica disponibilidade e gera código aleatório para a venda    
        produto = conn.execute("SELECT Nome, Valor, Quantidade FROM produtos WHERE Código=?", (codigo,))
        produto = produto.fetchone()
        if produto:
            nome, valor, quantidade = produto
            quantidade_vendida = int(janela_venda_combobox.get())
            if quantidade_vendida <= quantidade:
                if codigo_venda is None:
                    codigo_venda = random.randint(1000000, 9999999)
                lista_venda.append((codigo_venda, nome, valor, quantidade_vendida))
                janela_venda_label4.configure(text="Produto adicionado à lista!")
            else:
                janela_venda_label4.configure(text="Quantidade em estoque insuficiente.")
        else:
            janela_venda_label4.configure(text="Produto não encontrado.")

    # Adiciona manualmente um valor de pesagem à lista de vendas
    def adicionar_pesagem(valor):
        global codigo_venda
        if valor:
            valor = float(valor)
            if codigo_venda is None:
                codigo_venda = random.randint(1000000, 9999999)
            lista_venda.append((codigo_venda, "Pesagem", valor, 1))
            janela_venda_label4.configure(text="Produto adicionado à lista!")
        else:
            janela_venda_label4.configure(text="Insira um valor válido.")

    # Registra uma venda e atualiza o estoque de produtos
    def registrar_venda():
        global codigo_venda
        if not lista_venda:
            janela_venda_label4.configure(text="Lista de venda vazia.")
            return

        conexao = sqlite3.connect("Padaria.DB")
        now = datetime.now()
        data = now.date()
        hora = now.time().strftime("%H:%M:%S")
        total = 0
        for produto in lista_venda:
            codigo = codigo_venda
            nome = produto[1]
            valor = produto[2]
            quantidade_vendida = produto[3]
            conexao.execute("INSERT INTO vendas (Produtos, Valor, Quantidade, Código, Data, Hora) VALUES (?,?,?,?,?,?)", (nome, valor, quantidade_vendida, codigo, data, hora))
        total += valor * quantidade_vendida
        produto_db = conexao.execute("SELECT Quantidade FROM produtos WHERE Código=?", (codigo,)).fetchone()
        if produto_db:
            quantidade = produto_db[0]
            quantidade -= quantidade_vendida
            conexao.execute("UPDATE produtos SET Quantidade=? WHERE Código=?", (quantidade, codigo))
            conexao.commit()
            conexao.close()
            lista_venda.clear()
            janela_venda_label4.configure(text="Venda registrada com sucesso! Total: R$ {:.2f}".format(total))

    #Criação dos widgets da tela de venda
    janela_venda_label1 = Label(janela_venda, text= "Insira o código do produto", font=("Arial", 12), fg="red")
    janela_venda_label1.grid(column=0, row=1, padx=10, pady=30, sticky="NSEW", columnspan=5)

    janela_venda_entry1 = Entry(janela_venda, font=("Arial", 12), fg="Black", background= "beige")
    janela_venda_entry1.grid(column=0, row=2, padx=10, pady=10, sticky="NSEW", columnspan=5, ipadx=40, ipady=5)

    janela_venda_label2 = Label(janela_venda, text= "Selecione a quantidade a ser vendida", font=("Arial", 12), fg="red")
    janela_venda_label2.grid(column=8, row=1, padx=10, pady=10, sticky="NSEW", columnspan=5)

    janela_venda_combobox = ttk.Combobox(janela_venda, font=("Arial", 12),values=lista, state="readonly")
    janela_venda_combobox.grid(column=8, row=2, padx=10, pady=10, sticky="NSEW", columnspan=5, ipadx=40, ipady=5)
 
    janela_venda_button = Button(janela_venda, text="Guardar produto", font=("Arial", 12), fg="red", command=guardar_produto)
    janela_venda_button.grid(column=0, row=5, padx=10, pady=30, sticky="NSEW", columnspan=5)
  
    janela_venda_button2 = Button(janela_venda, text="Registrar venda", font=("Arial", 12), fg="red", command=registrar_venda)
    janela_venda_button2.grid(column=8, row=5, padx=10, pady=30, sticky="NSEW", columnspan=5)

    janela_venda_button3 = Button(janela_venda, text="Deletar", font=("Arial", 12), fg="red", command=limpar_lista)
    janela_venda_button3.grid(column=0, row=7, padx=10, pady=140, sticky="NSEW", columnspan=3)

    janela_venda_label4 = Label(janela_venda, font=("Arial", 12), fg="red")
    janela_venda_label4.grid(column=0, row=6, padx=10, pady=10, sticky="NSEW", columnspan=5)
    janela_venda_label4.place(relx=0.5, rely=0.5, anchor=CENTER)

def abrir_janela_registro(janela_inicial):
 #Criaçaõ da janela de registro
 janela_registro = Toplevel(janela_inicial)
 janela_registro.geometry("570x540") 
 janela_registro.resizable(False, False)
 janela_registro.title("Tela de registro")

 #Define a imagem de fundo da janela
 background_label = Label(janela_registro, image=background_image3)
 background_label.place(x=0, y=0, relwidth=1, relheight=1)
 def gerar_registro():
    # Abrindo conexão com o banco de dados
    conexao = sqlite3.connect("Padaria.DB")

    # Utilizando o with para garantir que a conexão será fechada corretamente, mesmo em caso de exceções
    with sqlite3.connect("Padaria.DB") as conexao:
        # Realizando uma busca no banco de dados pelo nome do produto a ser registrado
        produto = conexao.execute("SELECT * FROM produtos WHERE Nome = ?", (janela_registro_entry1.get(),)).fetchone()
        if produto:
            # Se o produto já existir, atualiza os valores do produto com o mesmo nome
            conexao.execute("UPDATE produtos SET Valor = ?, Quantidade = Quantidade + ? WHERE Nome = ?",
                             (janela_registro_entry2.get(), int(janela_registro_combobox.get()), janela_registro_entry1.get()))
        else:
            # Se o produto não existir, insere um novo registro na tabela
            conexao.execute("INSERT INTO produtos (Código, Nome, Valor, Quantidade) VALUES (NULL, ?, ?, ?)",
                             (janela_registro_entry1.get(), janela_registro_entry2.get(), int(janela_registro_combobox.get())))
        
        # Realizando o commit das alterações e fechando a conexão
        conexao.commit()
        conexao.close()

        # Exibindo uma mensagem de sucesso
        mostrar_mensagem_sucesso()

#Função responsável por exibir uma mensagem de sucesso na tela
 def mostrar_mensagem_sucesso():
    resultado_label.grid(row=6, column=1, columnspan=10, padx=20, pady=150, sticky="NSEW", ipadx=50)

 #Criação dos widgets da janela de registro
 janela_registro_label1 = Label(janela_registro, text= "Insira o nome do produto", font=("Arial", 12), fg="red")
 janela_registro_label1.grid(column=0, row=1, padx=10, pady=30, sticky="NSEW", columnspan=5)

 janela_registro_entry1 = Entry(janela_registro, font=("Arial", 12), fg="Black", background= "beige")
 janela_registro_entry1.grid(column=0, row=2, padx=10, pady=10, sticky="NSEW", columnspan=5, ipadx=40, ipady=5)

 janela_registro_label2 = Label(janela_registro, text= "Insira o valor do produto", font=("Arial", 12), fg="red")
 janela_registro_label2.grid(column=7, row=1, padx=10, pady=30, sticky="NSEW", columnspan=5)

 janela_registro_entry2 = Entry(janela_registro, font=("Arial", 12), fg="Black", background="beige")
 janela_registro_entry2.grid(column=7, row=2, padx=10, pady=10, sticky="NSEW", columnspan=5, ipadx=40, ipady=5)

 janela_registro_label3 = Label(janela_registro, text="Insira a quantidade", font=("Arial", 12), fg="red", background="white")
 janela_registro_label3.grid(column=3, row=3, padx=30, pady=10, sticky="NSEW", columnspan=6, ipadx=40, ipady=5)

 janela_registro_combobox = ttk.Combobox(janela_registro, values=lista, font=("Arial", 14))

 janela_registro_combobox.grid(column=3, row=4, padx=1, pady=13, sticky="NSEW", columnspan=6)

 janela_registro_button = Button(janela_registro, text= "Registrar produto", font=("Arial", 15), command= gerar_registro)
 janela_registro_button.grid(column=3, row=5, padx=0, pady=25, sticky="NSEW", columnspan=5, ipadx=50, ipady=8)
 janela_registro_button.place(relx=0.5, rely=0.5, anchor=CENTER)
 
 resultado_label = Label(janela_registro, text="Produto registrado com sucesso!", font=("Arial", 12), bg="gray", fg="white")
 
#Criação dos widgets da janela inicial e incluindo as funções nos botões
janela_inicial.title("Janela inicial")
janela_inicial.geometry("410x450")
janela_inicial.resizable(False, False)

Label(janela_inicial, image=background_image2).place(x=0, y=0, relwidth=1, relheight=1)

janela_inicial_label = Label(janela_inicial, text= "Selecione a operação desejada", font=("Arial", 14), fg="red")
janela_inicial_label.grid(column=0, row=0, padx=73, pady=10, sticky= "nswe", columnspan=12)

janela_inicial_button1 = Button(janela_inicial, text= "Operação de registro", font=("Arial", 14), fg="red", command=lambda: abrir_janela_registro(janela_inicial))
janela_inicial_button1.grid(column=0, row=1, padx=73, pady=30, sticky= "nswe", columnspan=12)

janela_inicial_button2 = Button(janela_inicial, text= "Consulta de estoque", font=("Arial", 14), fg="red", command=lambda: consulta_estoque())
janela_inicial_button2.grid(column=0, row=2, padx=73, pady=30, sticky= "nswe", columnspan=12)

janela_inicial_button3 = Button(janela_inicial, text= "Operação de venda", font=("Arial", 14), fg="red", command=lambda: abrir_janela_venda(janela_inicial))
janela_inicial_button3.grid(column=0, row=3, padx=73, pady=30, sticky= "nswe", columnspan=12)

janela_inicial_button4 = Button(janela_inicial, text= "Consulta de vendas", font=("Arial", 14), fg="red", command=lambda:janela_excel())
janela_inicial_button4.grid(column=0, row=4, padx=73, pady=30, sticky= "nswe", columnspan=12)

janela_inicial.mainloop()